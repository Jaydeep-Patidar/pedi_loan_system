from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.tokens import default_token_generator
from django.contrib import messages
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Sum, Q, Count
from django.http import HttpResponse
from django.middleware.csrf import get_token
from django.utils import timezone
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
import razorpay
import json
import random
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from django.urls import reverse

from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.core.exceptions import ValidationError

from .models import Member, Pedi, MemberPedi, Payment, Loan, LoanPayment, Transaction, LoanTransaction, LoanApplication, LoanApplicationSettings, Notice
from .forms import MemberForm, PediForm, LoanForm, PasswordResetRequestForm, SetPasswordForm, PasswordChangeForm, NoticeForm, ReactivateMemberForm
from withdrawals.models import WithdrawalRequest, Withdrawal
from .decorators import admin_required, member_required
from .authentication import generate_jwt_token, decode_jwt_token
from .utils import apply_search, apply_status_filter, apply_sorting, paginate_queryset
from .utils_payment import get_payment_breakdown, complete_payment
from .utils_loan_payment import get_loan_payment_breakdown, complete_loan_payment
from withdrawals.services import (
    can_member_withdraw,
    calculate_withdrawal_amount,
    get_admin_request_metrics,
    get_member_withdrawal_snapshot,
    has_open_withdrawal_request,
    has_pending_withdrawal,
)
from withdrawals.eligibility import can_member_request_exit
from pedi_loan_system.settings import RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET

from datetime import timedelta
from dateutil.relativedelta import relativedelta

available_fund = None


def calculate_available_balance():
    """Return the available balance for issuing or approving loans.

    This uses the configured external available_fund if provided, otherwise
    falls back to a calculation based on paid pedi collections, loan receipts,
    active loan dues, and completed withdrawals.
    """
    if available_fund:
        try:
            available_amount = available_fund()
            if available_amount is not None:
                return max(Decimal('0.00'), available_amount)
        except Exception:
            pass

    pedi_totals = Payment.aggregate_paid_totals()
    total_paid_pedi_collection = pedi_totals['total']
    total_loan_collection = Decimal('0.00')
    for lp in LoanPayment.objects.only('amount', 'base_amount_paid', 'penalty_paid', 'total_paid'):
        total_loan_collection += lp.get_total_collected()
    total_active_loan_due = Loan.objects.filter(status='Active').aggregate(total=Sum('remaining_due'))['total'] or Decimal('0.00')
    total_withdrawn = Withdrawal.objects.filter(status='Completed').aggregate(total=Sum('withdrawal_amount'))['total'] or Decimal('0.00')
    available_amount = total_paid_pedi_collection + total_loan_collection - total_active_loan_due - total_withdrawn
    return max(Decimal('0.00'), available_amount)


def _get_pedi_schedule_end(pedi, membership_end_date=None):
    """Return an exclusive end date for payment generation."""
    if pedi.end_date:
        end_date = pedi.end_date
    else:
        end_date = pedi.start_date + relativedelta(months=pedi.duration_months)

    if membership_end_date:
        membership_end_exclusive = membership_end_date + timedelta(days=1)
        return min(end_date, membership_end_exclusive)
    return end_date


def generate_member_payments(member, pedi, start_month, start_year, membership_end_date=None):
    """Generate missing monthly payments from the member's joining month onward."""
    start_date = date(start_year, start_month, 1)
    generation_end = _get_pedi_schedule_end(pedi, membership_end_date)
    if not generation_end or start_date >= generation_end:
        return []

    current_date = start_date
    payments = []
    while current_date < generation_end:
        due_day = min(pedi.monthly_due_day or 1, 28)
        due_date = date(current_date.year, current_date.month, due_day)
        payment, created = Payment.objects.get_or_create(
            member=member,
            pedi=pedi,
            month=current_date.month,
            year=current_date.year,
            defaults={
                'amount': pedi.monthly_amount,
                'status': 'Pending',
                'due_date': due_date,
                'penalty_enabled': pedi.penalty_enabled,
                'grace_days': pedi.grace_days,
                'enable_late_fee_per_day': pedi.enable_late_fee_per_day,
                'late_fee_per_day': pedi.late_fee_per_day,
                'enable_fixed_penalty': pedi.enable_fixed_penalty,
                'fixed_penalty_amount': pedi.fixed_penalty_amount,
                'enable_percentage_penalty': pedi.enable_percentage_penalty,
                'percentage_penalty_rate': pedi.percentage_penalty_rate,
            }
        )
        if not payment.due_date:
            payment.due_date = due_date
            payment.save(update_fields=['due_date'])
        payments.append(payment)
        current_date += relativedelta(months=1)

    return payments


# Initialize Razorpay client
client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

# ---------------------- Authentication ----------------------
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            # Check if member is active
            try:
                member = user.member_profile
                if not member.is_active:
                    messages.error(request, 'Your account has been deactivated. Please contact the administrator.')
                    return render(request, 'login.html')
            except:
                # If user doesn't have member profile, allow login (e.g., admin/staff)
                pass
            
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def token_obtain_pair(request):
    username = request.data.get('username')
    password = request.data.get('password')
    user = authenticate(request, username=username, password=password)
    if user is None:
        return Response({'detail': 'Invalid credentials'}, status=status.HTTP_400_BAD_REQUEST)

    # Check if member is active
    try:
        member = user.member_profile
        if not member.is_active:
            return Response({'detail': 'Your account has been deactivated. Please contact the administrator.'}, status=status.HTTP_403_FORBIDDEN)
    except:
        # If user doesn't have member profile, allow (e.g., admin/staff)
        pass

    access_token = generate_jwt_token(user, token_type='access')
    refresh_token = generate_jwt_token(user, token_type='refresh')
    return Response({'access': access_token, 'refresh': refresh_token})

@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def token_refresh(request):
    refresh_token = request.data.get('refresh')
    if not refresh_token:
        return Response({'detail': 'Refresh token required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        payload = decode_jwt_token(refresh_token, required_token_type='refresh')
    except Exception as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    user = get_user_model().objects.filter(pk=payload.get('user_id'), is_active=True).first()
    if not user:
        return Response({'detail': 'User not found or inactive'}, status=status.HTTP_400_BAD_REQUEST)

    access_token = generate_jwt_token(user, token_type='access')
    return Response({'access': access_token})


def password_reset_request(request):
    if request.method == 'POST':
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            users = get_user_model().objects.filter(email=email, is_active=True)
            for user in users:
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                token = default_token_generator.make_token(user)
                reset_url = request.build_absolute_uri(reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token}))
                subject = 'Pedi Loan System Password Reset'
                message = (
                    f'Hello {user.username},\n\n'
                    'You requested a password reset. Use the link below to set a new password:\n\n'
                    f'{reset_url}\n\n'
                    'If you did not request this, please ignore this message.\n'
                )
                send_mail(subject, message, None, [user.email], fail_silently=False)
            messages.success(request, 'If an account with that email exists, a password reset link has been sent.')
            return redirect('password_reset_done')
    else:
        form = PasswordResetRequestForm()
    return render(request, 'password_reset.html', {'form': form})


def password_reset_done(request):
    return render(request, 'password_reset_done.html')


def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = get_user_model().objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, get_user_model().DoesNotExist):
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        messages.error(request, 'The password reset link is invalid or has expired.')
        return redirect('password_reset_request')

    if request.method == 'POST':
        form = SetPasswordForm(request.POST)
        if form.is_valid():
            user.set_password(form.cleaned_data['new_password1'])
            user.save()
            messages.success(request, 'Your password has been reset successfully. Please log in with your new password.')
            return redirect('login')
    else:
        form = SetPasswordForm()

    return render(request, 'password_reset_confirm.html', {'form': form})


@login_required
def password_change(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.POST)
        if form.is_valid():
            if not request.user.check_password(form.cleaned_data['old_password']):
                messages.error(request, 'Old password is incorrect.')
            else:
                request.user.set_password(form.cleaned_data['new_password1'])
                request.user.save()
                messages.success(request, 'Your password has been changed successfully.')
                return redirect('dashboard')
    else:
        form = PasswordChangeForm()
    return render(request, 'password_change.html', {'form': form})

# ---------------------- Dashboard ----------------------
@login_required
def dashboard(request):
    # Ensure user has a Member profile (create if missing)
    member, created = Member.objects.get_or_create(
        user=request.user,
        defaults={
            'role': 'admin' if request.user.is_superuser else 'member',
            'phone': '',
            'address': ''
        }
    )
    if created and request.user.is_superuser:
        member.role = 'admin'
        member.save()
    
    # Check if member is inactive (applies only to non-admin members)
    if member.role == 'member' and not member.is_active:
        messages.error(request, 'Your account has been deactivated. You cannot access the system.')
        logout(request)
        return redirect('login')
    
    # Now redirect to the correct dashboard based on role
    if member.role == 'admin':
        return redirect('admin_dashboard')
    else:
        return redirect('member_dashboard')
@login_required
@admin_required
def admin_dashboard(request):
    total_members = Member.objects.filter(role='member', is_active=True).count()
    paid_pedi_qs = Payment.objects.filter(status='Paid', is_cancelled=False)
    pedi_totals = Payment.aggregate_paid_totals(paid_pedi_qs)
    total_contribution = pedi_totals['contribution']
    total_pedi_fine_collected = pedi_totals['fine']
    total_collection = pedi_totals['total']
    total_loan_collection = Decimal('0.00')
    total_loan_fine_collected = Decimal('0.00')
    for lp in LoanPayment.objects.only('amount', 'base_amount_paid', 'penalty_paid', 'total_paid'):
        total_loan_collection += lp.get_total_collected()
        total_loan_fine_collected += lp.get_penalty_collected()
    total_loans = Loan.objects.filter(status='Active').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    pending_dues = Loan.objects.filter(status='Active').aggregate(total=Sum('remaining_due'))['total'] or Decimal('0.00')

    total_loan_profit = Decimal('0.00')
    total_loan_fine_outstanding = Decimal('0.00')
    for loan in Loan.objects.all():
        profit = loan.paid_amount - loan.amount
        if profit > 0:
            total_loan_profit += profit
    for loan in Loan.objects.filter(status='Active'):
        total_loan_fine_outstanding += loan.get_outstanding_penalty()
    total_loan_fine = total_loan_fine_collected + total_loan_fine_outstanding

    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    overdue_payments_count = 0
    for payment in Payment.objects.filter(status='Pending', is_cancelled=False).select_related('pedi'):
        if (payment.year > current_year) or (payment.year == current_year and payment.month > current_month):
            continue
        if payment.get_effective_overdue_status(today=today) == 'Overdue':
            overdue_payments_count += 1

    grand_total_collection = total_collection + total_loan_collection

    monthly_summary = []
    for month in range(1, 13):
        month_payments = paid_pedi_qs.filter(year=current_year, month=month)
        amount = Payment.aggregate_paid_totals(month_payments)['total']
        monthly_summary.append({'month': month, 'amount': float(amount)})

    recent_payments = paid_pedi_qs.select_related('member', 'pedi').order_by('-payment_date')[:5]
    recent_members = Member.objects.filter(role='member', is_active=True).select_related('user').order_by('-joined_date')[:5]

    # Get all members with their payment information
    all_members_payments = []
    members = Member.objects.filter(role='member', is_active=True).select_related('user')
    for member in members:
        total_paid = member.total_paid
        active_loans = member.loans.filter(status='Active')
        total_loan_due = active_loans.aggregate(total=Sum('remaining_due'))['total'] or 0
        last_payment = member.payments.filter(status='Paid', is_cancelled=False).order_by('-payment_date').first()
        
        all_members_payments.append({
            'member': member,
            'total_paid': total_paid,
            'total_loan_due': total_loan_due,
            'last_payment_date': last_payment.payment_date if last_payment else None,
            'last_payment_amount': last_payment.get_collected_total() if last_payment else 0,
        })

    context = {
        'total_members': total_members,
        'total_collection': total_collection,
        'total_contribution': total_contribution,
        'total_pedi_fine_collected': total_pedi_fine_collected,
        'overdue_payments_count': overdue_payments_count,
        'total_loan_collection': total_loan_collection,
        'grand_total_collection': grand_total_collection,
        'total_loans': total_loans,
        'pending_dues': pending_dues,
        'total_loan_profit': total_loan_profit,
        'total_loan_fine': total_loan_fine,
        'total_loan_fine_collected': total_loan_fine_collected,
        'monthly_summary_json': json.dumps(monthly_summary),
        'recent_payments': recent_payments,
        'recent_members': recent_members,
        'all_members_payments': all_members_payments,
        'current_year': current_year,
        # Withdrawal & exit metrics
        'request_metrics': get_admin_request_metrics(),
        'pending_withdrawals_count': WithdrawalRequest.objects.filter(status='Pending').count(),
        'total_pending_withdrawable': WithdrawalRequest.objects.filter(status='Pending').aggregate(total=Sum('calculated_amount'))['total'] or Decimal('0.00'),
        'recent_withdrawals': Withdrawal.objects.filter(status='Completed').select_related('member').order_by('-processed_at')[:5],
        'pending_exit_requests': MemberPedi.objects.filter(status='Exit Requested').select_related('member__user', 'pedi')[:10],
        'pending_withdrawal_requests': [
            {'request': wr, 'snapshot': get_member_withdrawal_snapshot(wr.member)}
            for wr in WithdrawalRequest.objects.filter(status='Pending').select_related('member__user')[:10]
        ],
        'under_review_withdrawals': WithdrawalRequest.objects.filter(status='Approved').select_related('member__user')[:10],
    }
    return render(request, 'admin_dashboard.html', context)

@login_required
@member_required
def member_dashboard(request):
    member = request.user.member_profile
    total_paid = member.total_paid
    loans = member.loans.filter(status='Active')
    total_loan_due = loans.aggregate(total=Sum('remaining_due'))['total'] or 0
    payments = member.payments.filter(status='Paid', is_cancelled=False).order_by('-payment_date')[:10]

    # Withdrawal info for member
    withdrawal_snapshot = get_member_withdrawal_snapshot(member)
    has_pending = has_pending_withdrawal(member)
    has_open = has_open_withdrawal_request(member)

    context = {
        'member': member,
        'total_paid': total_paid,
        'total_loan_due': total_loan_due,
        'active_loans': loans,
        'recent_payments': payments,
        'withdrawable_amount': withdrawal_snapshot['withdrawable_amount'],
        'can_withdraw': withdrawal_snapshot['can_withdraw'],
        'withdraw_errors': withdrawal_snapshot['errors'],
        'has_pending_withdrawal': has_pending,
        'has_open_withdrawal': has_open,
        'withdrawal_snapshot': withdrawal_snapshot,
    }

    return render(request, 'member_dashboard.html', context)

# ---------------------- Member Management (Admin) ----------------------
@login_required
@admin_required
def member_list(request):
    members = Member.objects.filter(role='member').select_related('user')
    members, search_term = apply_search(request, members, [
        'user__first_name', 'user__last_name', 'user__username', 'user__email', 'phone'
    ])
    status = request.GET.get('status', '')
    if status == 'Active':
        members = members.filter(is_active=True)
    elif status == 'Inactive':
        members = members.filter(is_active=False)

    sort_map = {
        'name': 'user__first_name',
        'username': 'user__username',
        'joined': 'joined_date',
        'status': 'is_active',
    }
    members, sort_key, sort_dir = apply_sorting(request, members, sort_map, default_order='-joined_date')
    page_obj = paginate_queryset(request, members, default_per_page=10)
    return render(request, 'member_list.html', {
        'members': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'status': status,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def member_create(request):
    if request.method == 'POST':
        form = MemberForm(request.POST)
        if form.is_valid():
            member = form.save()
            messages.success(request, f'Member {member.user.username} created successfully')
            return redirect('member_list')
    else:
        form = MemberForm()
    return render(request, 'member_form.html', {'form': form, 'title': 'Add Member'})

@login_required
@admin_required
def member_edit(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        form = MemberForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, 'Member updated successfully')
            return redirect('member_list')
    else:
        form = MemberForm(instance=member)
    return render(request, 'member_form.html', {'form': form, 'title': 'Edit Member', 'member_obj': member})


@login_required
@admin_required
def member_activate(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        member.is_active = True
        member.user.is_active = True
        member.user.save()
        member.save()
        messages.success(
            request,
            'Member activated successfully. They can only be assigned to new pedis — closed/exited pedi memberships cannot be rejoined.',
        )
        return redirect('member_list')
    return render(request, 'confirm_delete.html', {'object': member, 'title': 'Activate Member', 'button_label': 'Yes, Activate'})

@login_required
@admin_required
def member_delete(request, pk):
    member = get_object_or_404(Member, pk=pk)
    # Soft-deactivate member instead of hard delete
    if request.method == 'POST':
        # Check deactivation restrictions
        has_active_loans = Loan.objects.filter(member=member, status='Active').exists()
        has_pending_pedi_dues = Payment.objects.filter(member=member, status='Pending', is_cancelled=False).exists()
        if has_active_loans or has_pending_pedi_dues:
            messages.error(request, 'Member cannot be deactivated because active loans or pending pedi dues exist.')
            return redirect('member_list')

        member.is_active = False
        member.user.is_active = False
        member.user.save()
        member.save()
        messages.success(request, 'Member deactivated successfully')
        return redirect('member_list')

    return render(request, 'confirm_delete.html', {
        'object': member,
        'title': 'Deactivate Member',
        'button_label': 'Yes, Deactivate'
    })

# ---------------------- Pedi Management ----------------------
@login_required
@admin_required
def pedi_list(request):
    pedis = Pedi.objects.annotate(member_count=Count('member_pedis'))
    pedis, search_term = apply_search(request, pedis, ['name'])
    status = request.GET.get('status', '')
    if status == 'Active':
        pedis = pedis.filter(is_active=True)
    elif status == 'Inactive':
        pedis = pedis.filter(is_active=False)

    sort_map = {
        'name': 'name',
        'start_date': 'start_date',
        'member_count': '-member_count',
    }
    pedis, sort_key, sort_dir = apply_sorting(request, pedis, sort_map, default_order='-created_at')
    page_obj = paginate_queryset(request, pedis, default_per_page=10)
    return render(request, 'pedi_list.html', {
        'pedis': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'status': status,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def pedi_create(request):
    if request.method == 'POST':
        form = PediForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pedi created successfully')
            return redirect('pedi_list')
    else:
        form = PediForm()
    return render(request, 'pedi_form.html', {'form': form, 'title': 'Create Pedi'})

@login_required
@admin_required
def pedi_edit(request, pk):
    pedi = get_object_or_404(Pedi, pk=pk)
    if request.method == 'POST':
        form = PediForm(request.POST, instance=pedi)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Pedi updated successfully')
                return redirect('pedi_list')
            except ValidationError as e:
                # Attach error to form non-field errors for display
                form.add_error(None, e.message)
    else:
        form = PediForm(instance=pedi)
    return render(request, 'pedi_form.html', {'form': form, 'title': 'Edit Pedi'})

@login_required
@admin_required
def assign_members(request, pedi_id):
    pedi = get_object_or_404(Pedi, pk=pedi_id)
    # Prevent assigning members to completed/closed pedis
    if getattr(pedi, 'pedi_status', 'Active') != 'Active':
        messages.error(request, 'Cannot assign members to a completed or closed pedi.')
        return redirect('pedi_list')
    members = Member.objects.filter(role='member', is_active=True)
    assigned = MemberPedi.objects.filter(pedi=pedi).values_list('member_id', flat=True)
    existing_memberships = {mp.member_id: mp for mp in MemberPedi.objects.filter(pedi=pedi)}

    if request.method == 'POST':
        selected_members = set(map(int, request.POST.getlist('members')))

        for member_id in selected_members:
            member = Member.objects.get(pk=member_id)
            member_pedi = existing_memberships.get(member_id)
            if member_pedi and member_pedi.status in ('Exited', 'Completed', 'Defaulted'):
                messages.warning(
                    request,
                    f'{member.user.get_full_name()} cannot rejoin a closed pedi ({pedi.name}). '
                    'Assign them only to a new pedi.',
                )
                continue
            if not member_pedi:
                member_pedi = MemberPedi.objects.create(
                    member=member,
                    pedi=pedi,
                    membership_start_date=timezone.now().date(),
                    status='Active',
                )
            elif member_pedi.status != 'Active':
                member_pedi.status = 'Active'
                member_pedi.membership_start_date = timezone.now().date()
                member_pedi.exit_date = None
                member_pedi.member_exit_requested_at = None
                member_pedi.member_exit_request_reason = ''
                member_pedi.save()
            if member_pedi.status == 'Active':
                generate_member_payments(
                    member=member,
                    pedi=pedi,
                    start_month=member_pedi.membership_start_date.month,
                    start_year=member_pedi.membership_start_date.year,
                    membership_end_date=member_pedi.membership_end_date,
                )

        messages.success(request, 'Members assigned successfully')
        return redirect('pedi_list')

    context = {'pedi': pedi, 'members': members, 'assigned': assigned, 'existing_memberships': existing_memberships}
    return render(request, 'assign_members.html', context)


@login_required
@admin_required
def pedi_member_exit(request, pedi_id, member_id):
    pedi = get_object_or_404(Pedi, pk=pedi_id)
    member = get_object_or_404(Member, pk=member_id)
    try:
        membership = MemberPedi.objects.get(member=member, pedi=pedi)
    except MemberPedi.DoesNotExist:
        messages.error(request, 'Membership not found')
        return redirect('pedi_list')

    if request.method == 'POST':
        exit_reason = request.POST.get('exit_reason', '')
        exit_date_str = request.POST.get('exit_date')
        if exit_date_str:
            try:
                exit_date = timezone.datetime.strptime(exit_date_str, '%Y-%m-%d').date()
            except Exception:
                exit_date = timezone.now().date()
        else:
            exit_date = timezone.now().date()

        membership.status = 'Exited'
        membership.exit_date = exit_date
        membership.admin_exit_at = timezone.now()
        membership.admin_exit_reason = exit_reason
        if not membership.exit_reason:
            membership.exit_reason = exit_reason
        membership.save()

        # Cancel future unpaid payments for this member+pedi
        exit_year = exit_date.year
        exit_month = exit_date.month
        future_payments = Payment.objects.filter(member=member, pedi=pedi, status='Pending', is_cancelled=False)
        for p in future_payments:
            if p.year > exit_year or (p.year == exit_year and p.month > exit_month):
                p.is_cancelled = True
                p.save()

        messages.success(request, 'Member exited from pedi and future payments cancelled')
        return redirect('assign_members', pedi_id=pedi.id)

    # GET: render admin approval form
    return render(request, 'member_exit.html', {
        'pedi': pedi,
        'member': member,
        'membership': membership,
        'form_title': 'Exit Member from Pedi',
        'submit_label': 'Exit Member and Cancel Future Payments',
    })


@login_required
@admin_required
def reject_member_exit_request(request, pedi_id, member_id):
    pedi = get_object_or_404(Pedi, pk=pedi_id)
    member = get_object_or_404(Member, pk=member_id)
    try:
        membership = MemberPedi.objects.get(member=member, pedi=pedi)
    except MemberPedi.DoesNotExist:
        messages.error(request, 'Membership not found')
        return redirect('pedi_list')

    if membership.status != 'Exit Requested':
        messages.error(request, 'There is no exit request to reject for this member.')
        return redirect('assign_members', pedi_id=pedi.id)

    membership.status = 'Active'
    membership.member_exit_requested_at = None
    membership.member_exit_request_reason = ''
    membership.save()

    messages.success(request, 'Exit request rejected and membership restored.')
    return redirect('assign_members', pedi_id=pedi.id)


@login_required
@member_required
def member_pedi_exit_request(request, pedi_id):
    member = request.user.member_profile
    pedi = get_object_or_404(Pedi, pk=pedi_id)
    try:
        membership = MemberPedi.objects.get(member=member, pedi=pedi)
    except MemberPedi.DoesNotExist:
        messages.error(request, 'Membership not found')
        return redirect('member_payments')

    if membership.status == 'Exited':
        messages.error(request, 'This pedi membership has already been exited.')
        return redirect('member_payments')

    can_exit, exit_errors = can_member_request_exit(member, pedi_id=pedi_id)
    if not can_exit and membership.status != 'Exit Requested':
        for err in exit_errors:
            messages.error(request, err)
        return redirect('member_payments')

    if request.method == 'POST':
        if has_open_withdrawal_request(member):
            messages.error(request, 'Cannot submit exit request while a withdrawal request is pending or under review.')
            return redirect('member_payments')
        can_exit, exit_errors = can_member_request_exit(member, pedi_id=pedi_id)
        if not can_exit:
            for err in exit_errors:
                messages.error(request, err)
            return redirect('member_payments')
        request_reason = request.POST.get('request_reason', '')
        membership.status = 'Exit Requested'
        membership.member_exit_requested_at = timezone.now()
        membership.member_exit_request_reason = request_reason
        membership.save()
        messages.success(request, 'Exit request submitted. An administrator will review it.')
        return redirect('member_payments')

    return render(request, 'member_exit.html', {
        'pedi': pedi,
        'member': member,
        'membership': membership,
        'form_title': 'Request Exit from Pedi',
        'submit_label': 'Request Exit',
        'request_mode': True,
    })

@login_required
@admin_required
def pedi_payment_history_menu(request):
    pedis = Pedi.objects.annotate(member_count=Count('member_pedis'))
    pedis, search_term = apply_search(request, pedis, ['name'], search_param='q')
    status = request.GET.get('status', '')
    if status == 'Active':
        pedis = pedis.filter(is_active=True)
    elif status == 'Inactive':
        pedis = pedis.filter(is_active=False)
    sort_map = {
        'name': 'name',
        'start_date': 'start_date',
        'members': '-member_count',
    }
    pedis, sort_key, sort_dir = apply_sorting(request, pedis, sort_map, default_order='-created_at')
    page_obj = paginate_queryset(request, pedis, default_per_page=10)
    return render(request, 'pedi_payment_history_menu.html', {
        'pedis': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'status': status,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def pedi_payment_history(request, pedi_id):
    selected_pedi = get_object_or_404(Pedi, pk=pedi_id)
    payments = Payment.objects.filter(pedi=selected_pedi, is_cancelled=False).select_related('member').order_by('-year', '-month', '-payment_date')
    payments, search_term = apply_search(request, payments, ['member__user__first_name', 'member__user__last_name', 'transaction_id'], search_param='q')
    sort_map = {
        'member': 'member__user__first_name',
        'date': '-payment_date',
        'amount': 'amount',
        'status': 'status',
    }
    payments, sort_key, sort_dir = apply_sorting(request, payments, sort_map, default_order='-year')
    page_obj = paginate_queryset(request, payments, default_per_page=10)
    return render(request, 'pedi_payment_history.html', {
        'selected_pedi': selected_pedi,
        'payments': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

# ---------------------- Monthly Payments ----------------------
@login_required
@admin_required
def monthly_payments(request, pedi_id=None):
    pedis = Pedi.objects.filter(is_active=True)
    selected_pedi = None
    payments = []

    # Accept pedi_id from either URL path or query parameters
    query_pedi_id = request.GET.get('pedi_id')
    if query_pedi_id and not pedi_id:
        try:
            pedi_id = int(query_pedi_id)
        except (TypeError, ValueError):
            pedi_id = None

    current_month = int(request.GET.get('month', timezone.now().month))
    current_year = int(request.GET.get('year', timezone.now().year))

    if pedi_id:
        selected_pedi = get_object_or_404(Pedi, pk=pedi_id)
        payments_qs = Payment.objects.filter(
            pedi=selected_pedi,
            month=current_month,
            year=current_year
        ).filter(is_cancelled=False).select_related('member').order_by('member__user__username')

        today = timezone.now().date()
        for payment in payments_qs:
            breakdown = get_payment_breakdown(payment, today=today)
            payments.append({
                'member': payment.member,
                'payment': payment,
                'amount': payment.amount,
                'status': payment.status,
                'breakdown': breakdown,
                'effective_status': breakdown['effective_status'],
            })

        if request.method == 'POST':
            if request.POST.get('confirm_payments') != '1':
                messages.error(request, 'Please confirm payment before marking as paid.')
                return redirect('monthly_payments', pedi_id=selected_pedi.id)

            selected_ids = []
            for item in payments:
                payment_id = item['payment'].id
                if request.POST.get(f'payment_{payment_id}'):
                    selected_ids.append(payment_id)

            if not selected_ids:
                messages.warning(request, 'No payments selected. Check "Mark as Paid" for at least one member.')
            else:
                updated_count = 0
                with transaction.atomic():
                    for payment_id in selected_ids:
                        payment = Payment.objects.select_for_update().get(pk=payment_id)
                        if payment.status == 'Paid':
                            continue
                        txn_id = payment.transaction_id
                        if not txn_id:
                            random_suffix = random.randint(1000, 9999)
                            txn_id = f'CASH-{timezone.now().strftime("%Y%m%d%H%M%S")}-{random_suffix}'
                        complete_payment(
                            payment,
                            payment_method='Cash',
                            transaction_id=txn_id,
                        )
                        updated_count += 1
                if updated_count:
                    messages.success(request, f'{updated_count} payment(s) marked as paid successfully.')
                else:
                    messages.info(request, 'Selected payments were already marked as paid.')

            return redirect('monthly_payments', pedi_id=selected_pedi.id)

    context = {
        'pedis': pedis,
        'selected_pedi': selected_pedi,
        'payments': payments,
        'current_month': current_month,
        'current_year': current_year,
        'months': range(1, 13),
    }
    return render(request, 'monthly_payments.html', context)

# ---------------------- Loan Management ----------------------
@login_required
@admin_required
def loan_list(request):
    loans = Loan.objects.all().select_related('member', 'member__user')
    loans, search_term = apply_search(request, loans, [
        'member__user__first_name', 'member__user__last_name', 'member__user__username'
    ])
    status = request.GET.get('status', '')
    if status:
        loans = loans.filter(status=status)

    sort_map = {
        'amount': 'amount',
        'issued': '-issued_date',
        'status': 'status',
        'member': 'member__user__first_name',
    }
    loans, sort_key, sort_dir = apply_sorting(request, loans, sort_map, default_order='-issued_date')
    page_obj = paginate_queryset(request, loans, default_per_page=10)
    return render(request, 'loan_list.html', {
        'loans': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'status': status,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def loan_create(request):
    available_amount = calculate_available_balance()

    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            requested_amount = form.cleaned_data['amount']
            if available_amount is not None and requested_amount > available_amount:
                form.add_error('amount', 'Insufficient collection available to issue this loan. Available amount: ₹{:.2f}'.format(available_amount))
            else:
                loan = form.save(commit=False)
                settings = LoanApplicationSettings.objects.filter(is_active=True).first()
                if not settings:
                    settings = LoanApplicationSettings.objects.order_by('-created_at').first()
                if not settings:
                    settings = SimpleNamespace(
                        default_interest_rate=Decimal('10.0'),
                        default_loan_duration_months=12,
                        penalty_enabled=False,
                        grace_days=0,
                        enable_late_fee_per_day=False,
                        late_fee_per_day=Decimal('0.00'),
                        enable_fixed_penalty=False,
                        fixed_penalty_amount=Decimal('0.00'),
                        enable_percentage_penalty=False,
                        percentage_penalty_rate=Decimal('0.00'),
                    )

                if loan.interest_rate is None:
                    loan.interest_rate = settings.default_interest_rate

                loan.due_date = timezone.now().date() + relativedelta(months=settings.default_loan_duration_months)
                loan.penalty_enabled = settings.penalty_enabled
                loan.grace_days = settings.grace_days
                loan.enable_late_fee_per_day = settings.enable_late_fee_per_day
                loan.late_fee_per_day = settings.late_fee_per_day
                loan.enable_fixed_penalty = settings.enable_fixed_penalty
                loan.fixed_penalty_amount = settings.fixed_penalty_amount
                loan.enable_percentage_penalty = settings.enable_percentage_penalty
                loan.percentage_penalty_rate = settings.percentage_penalty_rate
                loan.status = 'Active'
                loan.save()

                messages.success(request, 'Loan issued successfully')
                return redirect('loan_list')
    else:
        form = LoanForm()
    return render(request, 'loan_form.html', {
        'form': form,
        'title': 'Issue Loan',
        'available_amount': available_amount,
    })

@login_required
@admin_required
def loan_edit(request, pk):
    loan = get_object_or_404(Loan, pk=pk)
    if request.method == 'POST':
        form = LoanForm(request.POST, instance=loan)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Loan updated successfully')
                return redirect('loan_list')
            except ValidationError as e:
                form.add_error(None, e.message)
    else:
        form = LoanForm(instance=loan)
    return render(request, 'loan_form.html', {'form': form, 'title': 'Edit Loan'})

@login_required
@admin_required
def admin_loan_pay(request, loan_id):
    loan = get_object_or_404(Loan, pk=loan_id)
    if loan.status != 'Active':
        messages.warning(request, 'This loan is not active.')
        return redirect('loan_list')

    breakdown = get_loan_payment_breakdown(loan, loan.remaining_due or Decimal('0.00'))

    if request.method == 'POST':
        if request.POST.get('confirm_payment') != '1':
            messages.error(request, 'Please confirm payment after reviewing the penalty breakdown.')
            return redirect('admin_loan_pay', loan_id=loan.id)

        try:
            base_amount = Decimal(request.POST.get('amount', '0'))
        except Exception:
            messages.error(request, 'Please enter a valid amount.')
            return redirect('admin_loan_pay', loan_id=loan.id)

        preview = get_loan_payment_breakdown(loan, base_amount)
        if preview['final_payable'] <= 0:
            messages.error(request, 'Payment amount must be greater than zero.')
            return redirect('admin_loan_pay', loan_id=loan.id)

        remaining = loan.remaining_due or Decimal('0.00')
        if preview['base_amount'] > remaining and remaining > 0:
            messages.error(request, f'EMI amount cannot exceed remaining due ({remaining}).')
            return redirect('admin_loan_pay', loan_id=loan.id)

        random_suffix = random.randint(1000, 9999)
        payment, error = complete_loan_payment(
            loan,
            preview['base_amount'],
            payment_method='Cash',
            transaction_id=f'CASH-{timezone.now().strftime("%Y%m%d%H%M%S")}-{random_suffix}',
        )
        if error:
            messages.error(request, error)
            return redirect('admin_loan_pay', loan_id=loan.id)

        messages.success(
            request,
            f'Payment recorded: EMI {preview["base_amount"]} + penalty {preview["penalty_amount"]} '
            f'= {payment.get_total_collected()} for {loan.member.user.get_full_name()}.',
        )
        return redirect('loan_list')

    return render(request, 'admin_loan_pay.html', {
        'loan': loan,
        'breakdown': breakdown,
    })

# ---------------------- Member Views for Loans & Payments ----------------------
@login_required
@member_required
def member_loans(request):
    member = request.user.member_profile
    loans = member.loans.all()
    loans, search_term = apply_search(request, loans, [
        'amount', 'interest_rate', 'status', 'issued_date', 'due_date'
    ], search_param='q')
    sort_map = {
        'amount': 'amount',
        'due': 'remaining_due',
        'issue_date': '-issued_date',
        'status': 'status',
    }
    loans, sort_key, sort_dir = apply_sorting(request, loans, sort_map, default_order='-issued_date')
    page_obj = paginate_queryset(request, loans, default_per_page=10)
    return render(request, 'member_loans.html', {
        'loans': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@member_required
def member_payments(request):
    member = request.user.member_profile
    search_term = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year

    # Gather pedis the member has payments or memberships for
    memberships = MemberPedi.objects.filter(member=member)
    if status_filter == 'Active':
        memberships = memberships.filter(status='Active')
    elif status_filter == 'Exited':
        memberships = memberships.filter(status='Exited')
    elif status_filter == 'Closed':
        memberships = memberships.filter(status__in=['Completed', 'Defaulted'])

    pedi_ids = memberships.values_list('pedi_id', flat=True)
    pedis = Pedi.objects.filter(id__in=pedi_ids).distinct()
    if search_term:
        pedis = pedis.filter(name__icontains=search_term)
    pedi_sections = []
    for pedi in pedis:
        membership = MemberPedi.objects.filter(member=member, pedi=pedi).first()
        payments_qs = Payment.objects.filter(member=member, pedi=pedi, is_cancelled=False)

        # Only include payments that are Paid OR due/overdue, but NEVER show future months.
        visible_payments = []

        total_paid = Decimal('0.00')

        pending_amount = Decimal('0.00')
        total_penalty = Decimal('0.00')
        remaining_payments = 0

        for p in payments_qs.order_by('-year', '-month'):
            # Never show future months (even if due_date is after today).
            if (p.year > current_year) or (p.year == current_year and p.month > current_month):
                continue

            breakdown = get_payment_breakdown(p, today=today)
            penalty = breakdown['penalty_amount']
            final_payable = breakdown['final_payable']
            grace_days = breakdown['grace_days']

            if p.status == 'Paid':
                total_paid += p.get_collected_total()
                visible_payments.append({
                    'payment': p,
                    'badge': 'Paid',
                    'badge_class': 'bg-success',
                    'penalty': penalty,
                    'final_payable': final_payable,
                    'grace_days': grace_days,
                    'overdue_days': 0,
                })
            else:
                effective_status = breakdown['effective_status']
                badge = 'Overdue' if effective_status == 'Overdue' else 'Pending'
                badge_class = 'bg-danger' if badge == 'Overdue' else 'bg-warning'
                visible_payments.append({
                    'payment': p,
                    'badge': badge,
                    'badge_class': badge_class,
                    'overdue_days': breakdown['overdue_days'],
                    'penalty': penalty,
                    'final_payable': final_payable,
                    'grace_days': grace_days,
                    'status_effective': effective_status,
                })
                pending_amount += p.amount
                total_penalty += penalty
                remaining_payments += 1


        if membership:
            pedi_sections.append({
                'pedi': pedi,
                'membership': membership,
                'payments': visible_payments,
                'summary': {
                    'total_paid': total_paid,
                    'pending_amount': pending_amount,
                    'total_penalty': total_penalty,
                    'remaining_payments': remaining_payments,
                }
            })

    return render(request, 'member_payments.html', {
        'pedi_sections': pedi_sections,
        'search_term': search_term,
        'status': status_filter,
    })

@login_required
@member_required
def payment_history(request):
    member = request.user.member_profile
    payments = member.payments.filter(status='Paid', is_cancelled=False).order_by('-payment_date')
    payments, search_term = apply_search(request, payments, [
        'pedi__name', 'transaction_id', 'month', 'year'
    ], search_param='q')
    sort_map = {
        'date': '-payment_date',
        'amount': 'amount',
        'pedi': 'pedi__name',
    }
    payments, sort_key, sort_dir = apply_sorting(request, payments, sort_map, default_order='-payment_date')
    page_obj = paginate_queryset(request, payments, default_per_page=10)
    return render(request, 'payment_history.html', {
        'payments': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

# ---------------------- loan Payment ----------------------
@login_required
@member_required
def loan_pay_online(request, loan_id):
    loan = get_object_or_404(Loan, pk=loan_id, member=request.user.member_profile)
    if loan.status == 'Closed':
        messages.warning(request, 'This loan is already closed.')
        return redirect('member_loans')

    default_base = loan.remaining_due or Decimal('0.00')
    default_breakdown = get_loan_payment_breakdown(loan, default_base)

    if request.method == 'POST':
        try:
            base_amount = Decimal(request.POST.get('amount', '0'))
        except Exception:
            messages.error(request, 'Please enter a valid amount.')
            return redirect('loan_pay_online', loan_id=loan.id)

        breakdown = get_loan_payment_breakdown(loan, base_amount)
        if breakdown['final_payable'] <= 0:
            messages.error(request, 'Payment amount must be greater than zero.')
            return redirect('loan_pay_online', loan_id=loan.id)

        remaining = loan.remaining_due or Decimal('0.00')
        if breakdown['base_amount'] > remaining and remaining > 0:
            messages.error(request, f'EMI amount cannot exceed remaining due ({remaining}).')
            return redirect('loan_pay_online', loan_id=loan.id)

        order_amount = int(breakdown['final_payable'] * 100)
        try:
            razorpay_order = client.order.create({
                'amount': order_amount,
                'currency': 'INR',
                'payment_capture': '1',
                'notes': {
                    'loan_id': loan.id,
                    'member_id': loan.member.id,
                    'base_amount': str(breakdown['base_amount']),
                    'penalty_amount': str(breakdown['penalty_amount']),
                    'total_payable': str(breakdown['final_payable']),
                },
            })
        except Exception as e:
            messages.error(request, f'Payment gateway error: {str(e)}')
            return redirect('member_loans')

        LoanTransaction.objects.create(
            loan=loan,
            amount=breakdown['final_payable'],
            base_amount=breakdown['base_amount'],
            penalty_amount=breakdown['penalty_amount'],
            razorpay_order_id=razorpay_order['id'],
            status='Created',
        )

        payment_data = {
            'razorpay_key_id': RAZORPAY_KEY_ID,
            'amount_paise': order_amount,
            'razorpay_order_id': razorpay_order['id'],
            'csrf_token': get_token(request),
            'success_url': request.build_absolute_uri(reverse('loan_payment_online_success')),
            'base_amount': str(breakdown['base_amount']),
            'penalty_amount': str(breakdown['penalty_amount']),
            'total_payable': str(breakdown['final_payable']),
        }
        return render(request, 'loan_payment_gateway.html', {
            'loan': loan,
            'loan_payment_data': payment_data,
            'breakdown': breakdown,
        })

    return render(request, 'loan_pay_online.html', {
        'loan': loan,
        'breakdown': default_breakdown,
    })

@login_required
def loan_payment_online_success(request):
    if request.method == 'POST':
        razorpay_payment_id = request.POST.get('razorpay_payment_id')
        razorpay_order_id = request.POST.get('razorpay_order_id')
        razorpay_signature = request.POST.get('razorpay_signature')

        if not all([razorpay_payment_id, razorpay_order_id, razorpay_signature]):
            messages.error(request, 'Payment verification failed: missing payment details.')
            return redirect('member_loans')

        params_dict = {
            'razorpay_order_id': razorpay_order_id,
            'razorpay_payment_id': razorpay_payment_id,
            'razorpay_signature': razorpay_signature,
        }

        try:
            client.utility.verify_payment_signature(params_dict)
        except Exception:
            messages.error(request, 'Payment verification failed: invalid signature.')
            return redirect('member_loans')

        try:
            with transaction.atomic():
                loan_trans = LoanTransaction.objects.select_for_update().get(razorpay_order_id=razorpay_order_id)
                if hasattr(request.user, 'member_profile'):
                    if loan_trans.loan.member_id != request.user.member_profile.id:
                        messages.error(request, 'Payment verification failed: unauthorized.')
                        return redirect('member_loans')

                if loan_trans.status == 'Success':
                    messages.success(request, 'Loan payment already recorded successfully.')
                    return redirect('member_loans')

                loan_trans.razorpay_payment_id = razorpay_payment_id
                loan_trans.razorpay_signature = razorpay_signature
                loan_trans.status = 'Success'
                loan_trans.save(update_fields=['razorpay_payment_id', 'razorpay_signature', 'status'])

                base_amount = loan_trans.base_amount or Decimal('0.00')
                expected = (base_amount + loan_trans.penalty_amount).quantize(Decimal('0.01'))
                if loan_trans.amount != expected:
                    raise ValueError('Transaction amount does not match expected payable total.')

                payment, error = complete_loan_payment(
                    loan_trans.loan,
                    base_amount,
                    payment_method='Online',
                    transaction_id=razorpay_payment_id,
                )
                if error:
                    raise ValueError(error)

            messages.success(
                request,
                f'Loan payment successful! Total paid: ₹{payment.get_total_collected()} '
                f'(EMI ₹{payment.get_base_collected()} + penalty ₹{payment.get_penalty_collected()}).',
            )
            return redirect('member_loans')
        except LoanTransaction.DoesNotExist:
            messages.error(request, 'Payment verification failed: transaction not found.')
        except ValueError as exc:
            messages.error(request, f'Payment verification failed: {exc}')
        except Exception:
            messages.error(request, 'Payment verification failed. Contact support if amount was deducted.')
        return redirect('member_loans')
    return redirect('member_loans')

@login_required
@member_required
def loan_payment_history(request):
    member = request.user.member_profile
    payments = LoanPayment.objects.filter(loan__member=member).select_related('loan')
    payments, search_term = apply_search(request, payments, ['loan__id', 'loan__member__user__username'], search_param='q')
    sort_map = {
        'amount': 'amount',
        'date': '-payment_date',
        'method': 'payment_method',
    }
    payments, sort_key, sort_dir = apply_sorting(request, payments, sort_map, default_order='-payment_date')
    page_obj = paginate_queryset(request, payments, default_per_page=10)
    return render(request, 'loan_payment_history.html', {
        'payments': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def admin_loan_payments(request):
    payments = LoanPayment.objects.select_related('loan__member__user').order_by('-payment_date')
    payments, search_term = apply_search(request, payments, [
        'loan__member__user__first_name', 'loan__member__user__last_name', 'transaction_id'
    ], search_param='q')
    sort_map = {
        'amount': 'amount',
        'date': '-payment_date',
        'member': 'loan__member__user__first_name',
    }
    payments, sort_key, sort_dir = apply_sorting(request, payments, sort_map, default_order='-payment_date')
    page_obj = paginate_queryset(request, payments, default_per_page=10)
    return render(request, 'admin_loan_payments.html', {
        'payments': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })
# ---------------------- Online Payment ----------------------

@login_required
@member_required
def make_payment(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id, member=request.user.member_profile)
    if payment.status == 'Paid':
        messages.warning(request, 'This payment is already completed')
        return redirect('member_payments')

    if request.method == 'POST':
        # Create Razorpay Order (amount in paise)
        penalty_amount = payment.calculate_penalty()
        total_payable = (payment.amount + penalty_amount).quantize(Decimal('0.01'))
        order_amount = int(total_payable * 100)  # convert to paise (integer)
        order_currency = 'INR'
        razorpay_order = client.order.create({
            'amount': order_amount,
            'currency': order_currency,
            'payment_capture': '1',
            'notes': {
                'payment_id': payment.id,
                'member_id': payment.member.id,
                'penalty_amount': str(penalty_amount),
                'base_amount': str(payment.amount),
                'total_payable': str(total_payable),
            }
        })


        # Save transaction
        Transaction.objects.create(
            member=payment.member,
            payment=payment,
            razorpay_order_id=razorpay_order['id'],
            # Store the actual paid total (base + frozen penalty)
            amount=total_payable,
            status='Created'
        )


        payment_data = {
            'razorpay_key_id': RAZORPAY_KEY_ID,
            'amount_paise': order_amount,
            'razorpay_order_id': razorpay_order['id'],
            'csrf_token': get_token(request),
            'success_url': request.build_absolute_uri(reverse('payment_success')),
            'pedi_name': payment.pedi.name if payment.pedi else '',
            'month': payment.month,
            'year': payment.year,
            'base_amount': str(payment.amount),
            'penalty_amount': str(penalty_amount),
            'total_payable': str(total_payable),
        }

        context = {
            'payment': payment,
            'payment_data_json': payment_data,
            'breakdown': get_payment_breakdown(payment),
            'razorpay_order_id': razorpay_order['id'],
            'razorpay_key_id': RAZORPAY_KEY_ID,
            'amount': payment.amount,
            'amount_paise': order_amount,
        }
        return render(request, 'payment_gateway.html', context)

    breakdown = get_payment_breakdown(payment)
    return render(request, 'make_payment.html', {
        'payment': payment,
        'breakdown': breakdown,
    })

@login_required
def payment_success(request):
    if request.method == 'POST':
        razorpay_payment_id = request.POST.get('razorpay_payment_id')
        razorpay_order_id = request.POST.get('razorpay_order_id')
        razorpay_signature = request.POST.get('razorpay_signature')

        if not all([razorpay_payment_id, razorpay_order_id, razorpay_signature]):
            messages.error(request, 'Payment verification failed: missing payment details.')
            return redirect('member_payments')

        params_dict = {
            'razorpay_order_id': razorpay_order_id,
            'razorpay_payment_id': razorpay_payment_id,
            'razorpay_signature': razorpay_signature,
        }

        try:
            client.utility.verify_payment_signature(params_dict)
        except Exception:
            messages.error(request, 'Payment verification failed: invalid signature.')
            return redirect('member_payments')

        try:
            with transaction.atomic():
                txn = Transaction.objects.select_for_update().get(razorpay_order_id=razorpay_order_id)
                if not hasattr(request.user, 'member_profile') or txn.member_id != request.user.member_profile.id:
                    messages.error(request, 'Payment verification failed: unauthorized transaction.')
                    return redirect('member_payments')

                payment = txn.payment
                if not payment:
                    messages.error(request, 'Payment verification failed: payment record not found.')
                    return redirect('member_payments')

                if txn.status == 'Success' and payment.status == 'Paid':
                    messages.success(request, 'Payment already recorded successfully.')
                    return redirect('member_payments')

                txn.razorpay_payment_id = razorpay_payment_id
                txn.razorpay_signature = razorpay_signature
                txn.status = 'Success'
                txn.save(update_fields=['razorpay_payment_id', 'razorpay_signature', 'status'])

                if payment.status != 'Paid':
                    expected_total = (payment.amount + payment.calculate_penalty()).quantize(Decimal('0.01'))
                    if txn.amount != expected_total:
                        raise ValueError('Paid amount does not match expected payable total.')

                    complete_payment(
                        payment,
                        payment_method='Online',
                        transaction_id=razorpay_payment_id,
                        razorpay_order_id=razorpay_order_id,
                        razorpay_payment_id=razorpay_payment_id,
                    )

            messages.success(request, 'Payment successful!')
            return redirect('member_payments')
        except Transaction.DoesNotExist:
            messages.error(request, 'Payment verification failed: transaction not found.')
        except ValueError as exc:
            messages.error(request, f'Payment verification failed: {exc}')
        except Exception:
            messages.error(request, 'Payment verification failed. Please contact support if amount was deducted.')
        return redirect('member_payments')

    return redirect('member_dashboard')

# ---------------------- Reports & Export ----------------------
@login_required
@admin_required
def reports(request):
    return render(request, 'reports.html')

@login_required
@admin_required
def export_members_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    headers = ['Name', 'Username', 'Phone', 'Address', 'Joined Date', 'Total Paid']
    ws.append(headers)
    members = Member.objects.filter(role='member')
    for member in members:
        ws.append([
            member.user.get_full_name(),
            member.user.username,
            member.phone,
            member.address,
            member.joined_date.strftime('%Y-%m-%d'),
            float(member.total_paid)
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    wb.save(response)
    return response

@login_required
@admin_required
def export_payments_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = [
        'Member', 'Pedi', 'Month', 'Year', 'Base Amount', 'Penalty Paid', 'Total Paid',
        'Due Date', 'Grace Days', 'Status', 'Payment Date', 'Method', 'Transaction ID',
    ]
    ws.append(headers)
    payments = Payment.objects.select_related('member', 'pedi').all()
    for payment in payments:
        ws.append([
            payment.member.user.get_full_name(),
            payment.pedi.name,
            payment.month,
            payment.year,
            float(payment.get_contribution_collected() if payment.status == 'Paid' else payment.amount),
            float(payment.penalty_paid or 0),
            float(payment.get_collected_total() if payment.status == 'Paid' else payment.get_final_payable_amount()),
            payment.get_due_date_exact().strftime('%Y-%m-%d'),
            payment.get_display_grace_days(),
            payment.status,
            payment.payment_date.strftime('%Y-%m-%d %H:%M') if payment.payment_date else '',
            payment.payment_method,
            payment.transaction_id or '',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
    wb.save(response)
    return response

@login_required
@admin_required
def export_loans_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Loans"
    headers = ['Member', 'Amount', 'Interest Rate', 'Total Payable', 'Paid Amount', 'Remaining Due', 'Status']
    ws.append(headers)
    loans = Loan.objects.select_related('member').all()
    for loan in loans:
        ws.append([
            loan.member.user.get_full_name(),
            float(loan.amount),
            float(loan.interest_rate),
            float(loan.total_payable),
            float(loan.paid_amount),
            float(loan.remaining_due),
            loan.status
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=loans.xlsx'
    wb.save(response)
    return response


@login_required
@member_required
def apply_loan(request):
    settings = LoanApplicationSettings.objects.filter(is_active=True).first()
    if not settings:
        settings = LoanApplicationSettings.objects.order_by('-created_at').first()
    if not settings:
        messages.error(request, 'Loan application period not configured.')
        return redirect('member_dashboard')

    today = timezone.now().date()
    if today < settings.start_date or today > settings.end_date:
        messages.warning(request, 'Loan applications are currently closed.')
        return redirect('member_dashboard')

    # --- NEW: Check for existing active loan ---
    member = request.user.member_profile
    if Loan.objects.filter(member=member, status='Active').exists():
        messages.warning(request, 'You already have an active loan. You cannot apply for a new loan until the current loan is fully paid.')
        return redirect('member_dashboard')
    # -------------------------------------------

    # Check if member already has a pending application
    if LoanApplication.objects.filter(member=member, status='Pending').exists():
        messages.error(request, 'You already have a pending loan application.')
        return redirect('member_dashboard')

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', 0))
        purpose = request.POST.get('purpose', '')
        if amount <= 0:
            messages.error(request, 'Please enter a valid amount.')
            return render(request, 'apply_loan.html', {'settings': settings})

        LoanApplication.objects.create(
            member=member,
            requested_amount=amount,
            purpose=purpose,
            status='Pending'
        )
        messages.success(request, 'Your loan application has been submitted.')
        return redirect('member_dashboard')

    return render(request, 'apply_loan.html', {'settings': settings})

@login_required
@admin_required
def admin_loan_applications(request):
    applications = LoanApplication.objects.select_related('member__user').order_by('-applied_date')
    applications, search_term = apply_search(request, applications, [
        'member__user__first_name', 'member__user__last_name', 'member__user__username'
    ], search_param='q')
    status = request.GET.get('status', '')
    if status:
        applications = applications.filter(status=status)
    sort_map = {
        'amount': 'requested_amount',
        'applied': '-applied_date',
        'status': 'status',
    }
    applications, sort_key, sort_dir = apply_sorting(request, applications, sort_map, default_order='-applied_date')
    page_obj = paginate_queryset(request, applications, default_per_page=10)
    return render(request, 'admin_loan_applications.html', {
        'applications': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'status': status,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def approve_loan_application(request, pk):
    application = get_object_or_404(LoanApplication, pk=pk)
    if application.status != 'Pending':
        messages.warning(request, 'This application is no longer pending.')
        return redirect('admin_loan_applications')

    settings = LoanApplicationSettings.objects.filter(is_active=True).first()
    if not settings:
        settings = LoanApplicationSettings.objects.order_by('-created_at').first()
    if not settings:
        messages.error(request, 'Loan settings are not configured. Please set active loan settings before approving applications.')
        return redirect('admin_loan_applications')

    if request.method == 'POST':
        interest_rate = Decimal(request.POST.get('interest_rate', settings.default_interest_rate))
        due_date = request.POST.get('due_date')
        if not due_date:
            due_date = (timezone.now().date() + relativedelta(months=settings.default_loan_duration_months))
        else:
            from datetime import datetime
            due_date = datetime.strptime(due_date, '%Y-%m-%d').date()

        fund = calculate_available_balance()
        if application.requested_amount > fund:
            messages.error(request, f'Cannot approve: insufficient available fund (Available: ₹{fund}).')
            return redirect('admin_loan_applications')

        # Create actual Loan with penalty settings preserved from current loan settings
        loan = Loan.objects.create(
            member=application.member,
            amount=application.requested_amount,
            interest_rate=interest_rate,
            due_date=due_date,
            penalty_enabled=settings.penalty_enabled,
            grace_days=settings.grace_days,
            enable_late_fee_per_day=settings.enable_late_fee_per_day,
            late_fee_per_day=settings.late_fee_per_day,
            enable_fixed_penalty=settings.enable_fixed_penalty,
            fixed_penalty_amount=settings.fixed_penalty_amount,
            enable_percentage_penalty=settings.enable_percentage_penalty,
            percentage_penalty_rate=settings.percentage_penalty_rate,
            status='Active'
        )
        # Update application and link to the created loan
        application.loan = loan
        application.status = 'Approved'
        application.approved_date = timezone.now()
        application.approved_interest_rate = interest_rate
        application.approved_due_date = due_date
        application.admin_remarks = request.POST.get('remarks', '')
        application.save()

        messages.success(request, f'Loan application approved. Loan #{loan.id} created.')
        return redirect('admin_loan_applications')

    # Pre-fill form with defaults
    default_due = timezone.now().date() + relativedelta(months=settings.default_loan_duration_months)
    context = {
        'application': application,
        'default_interest': settings.default_interest_rate,
        'default_due': default_due,
    }
    return render(request, 'approve_loan_application.html', context)

@login_required
@admin_required
def reject_loan_application(request, pk):
    application = get_object_or_404(LoanApplication, pk=pk)
    if application.status != 'Pending':
        messages.warning(request, 'This application is no longer pending.')
        return redirect('admin_loan_applications')
    application.status = 'Rejected'
    application.save()
    messages.success(request, 'Loan application rejected.')
    return redirect('admin_loan_applications')

# admin loan setting view
@login_required
@admin_required
def admin_loan_settings(request):
    settings = LoanApplicationSettings.objects.filter(is_active=True).first()
    if not settings:
        settings = LoanApplicationSettings.objects.order_by('-created_at').first()
    if not settings:
        settings = LoanApplicationSettings.objects.create(
            start_date=timezone.now().date(),
            end_date=timezone.now().date() + timedelta(days=30),
            default_interest_rate=Decimal('10.0'),
            default_loan_duration_months=12,
            is_active=True,
        )

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        interest_rate = request.POST.get('default_interest_rate')
        duration = request.POST.get('default_loan_duration_months')
        if start_date and end_date and interest_rate and duration:
            from datetime import datetime
            LoanApplicationSettings.objects.filter(is_active=True).update(is_active=False)

            new_settings = LoanApplicationSettings(
                start_date=datetime.strptime(start_date, '%Y-%m-%d').date(),
                end_date=datetime.strptime(end_date, '%Y-%m-%d').date(),
                default_interest_rate=Decimal(interest_rate),
                default_loan_duration_months=int(duration),
                penalty_enabled='penalty_enabled' in request.POST,
                grace_days=int(request.POST.get('grace_days') or 0),
                enable_late_fee_per_day='enable_late_fee_per_day' in request.POST,
                late_fee_per_day=Decimal(request.POST.get('late_fee_per_day') or 0),
                enable_fixed_penalty='enable_fixed_penalty' in request.POST,
                fixed_penalty_amount=Decimal(request.POST.get('fixed_penalty_amount') or 0),
                enable_percentage_penalty='enable_percentage_penalty' in request.POST,
                percentage_penalty_rate=Decimal(request.POST.get('percentage_penalty_rate') or 0),
                is_active=True,
            )
            try:
                new_settings.full_clean()
                new_settings.save()
                messages.success(request, 'Loan application settings saved as a new active version.')
            except ValidationError as exc:
                error_message = exc.messages[0] if exc.messages else 'Please fix the penalty settings.'
                messages.error(request, error_message)
                return redirect('admin_loan_settings')
            return redirect('admin_loan_settings')
        else:
            messages.error(request, 'Please fill all required fields.')
            return redirect('admin_loan_settings')

    context = {'settings': settings}
    return render(request, 'admin_loan_settings.html', context)


@login_required
@admin_required
def admin_loan_settings_history(request):
    settings_history = LoanApplicationSettings.objects.order_by('-created_at')
    page_obj = paginate_queryset(request, settings_history, default_per_page=10)
    return render(request, 'admin_loan_settings_history.html', {
        'settings_history': page_obj,
        'page_obj': page_obj,
    })


# ---------------------- Notice Management ----------------------
@login_required
@admin_required
def notice_list(request):
    notices = Notice.objects.select_related('author').all().order_by('-created_at')
    notices, search_term = apply_search(request, notices, ['title', 'content', 'author__username'])
    status = request.GET.get('status', '')
    if status == 'Active':
        notices = notices.filter(is_active=True)
    elif status == 'Inactive':
        notices = notices.filter(is_active=False)

    sort_map = {
        'title': 'title',
        'created': '-created_at',
        'author': 'author__username',
    }
    notices, sort_key, sort_dir = apply_sorting(request, notices, sort_map, default_order='-created_at')
    page_obj = paginate_queryset(request, notices, default_per_page=10)
    return render(request, 'notice_list.html', {
        'notices': page_obj,
        'page_obj': page_obj,
        'search_term': search_term,
        'status': status,
        'sort_key': sort_key,
        'sort_dir': sort_dir,
    })

@login_required
@admin_required
def notice_create(request):
    if request.method == 'POST':
        form = NoticeForm(request.POST)
        if form.is_valid():
            notice = form.save(commit=False)
            notice.author = request.user
            notice.save()
            messages.success(request, 'Notice created successfully')
            return redirect('notice_list')
    else:
        form = NoticeForm()
    return render(request, 'notice_form.html', {'form': form, 'title': 'Create Notice'})

@login_required
@admin_required
def notice_edit(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    if request.method == 'POST':
        form = NoticeForm(request.POST, instance=notice)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notice updated successfully')
            return redirect('notice_list')
    else:
        form = NoticeForm(instance=notice)
    return render(request, 'notice_form.html', {'form': form, 'title': 'Edit Notice'})

@login_required
@admin_required
def notice_delete(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    if request.method == 'POST':
        notice.delete()
        messages.success(request, 'Notice deleted successfully')
        return redirect('notice_list')
    return render(request, 'confirm_delete.html', {'object': notice})

@login_required
@member_required
def member_notices(request):
    notices = Notice.objects.filter(is_active=True).order_by('-created_at')
    return render(request, 'member_notices.html', {'notices': notices})
