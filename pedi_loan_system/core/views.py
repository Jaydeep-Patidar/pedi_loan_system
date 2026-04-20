from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.middleware.csrf import get_token
from django.utils import timezone
import razorpay
import json
from datetime import date
from decimal import Decimal
from django.urls import reverse

from .models import Member, Pedi, MemberPedi, Payment, Loan, LoanPayment, Transaction, LoanTransaction, LoanApplication, LoanApplicationSettings
from .forms import MemberForm, PediForm, LoanForm
from .decorators import admin_required, member_required
from pedi_loan_system.settings import RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET

from datetime import timedelta
from dateutil.relativedelta import relativedelta


# Initialize Razorpay client
client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

# ---------------------- Authentication ----------------------
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

# ---------------------- Dashboard ----------------------
@login_required
def dashboard(request):
    # Ensure user has a Member profile (create if missing)
    member, created = Member.objects.get_or_create(
        user=request.user,
        defaults={
            'role': 'admin' if request.user.is_superuser else 'member',
            'phone': '',
            'address': ''
        }
    )
    if created and request.user.is_superuser:
        member.role = 'admin'
        member.save()
    
    # Now redirect to the correct dashboard based on role
    if member.role == 'admin':
        return redirect('admin_dashboard')
    else:
        return redirect('member_dashboard')
@login_required
@admin_required
def admin_dashboard(request):
    total_members = Member.objects.filter(role='member', is_active=True).count()
    total_collection = Payment.objects.filter(status='Paid').aggregate(total=Sum('amount'))['total'] or 0
    total_loans = Loan.objects.filter(status='Active').aggregate(total=Sum('amount'))['total'] or 0
    pending_dues = Loan.objects.filter(status='Active').aggregate(total=Sum('remaining_due'))['total'] or 0

    current_year = timezone.now().year
    monthly_summary = []
    for month in range(1, 13):
        amount = Payment.objects.filter(year=current_year, month=month, status='Paid').aggregate(total=Sum('amount'))['total'] or 0
        monthly_summary.append({'month': month, 'amount': float(amount)})

    recent_payments = Payment.objects.filter(status='Paid').select_related('member', 'pedi').order_by('-payment_date')[:10]

    context = {
        'total_members': total_members,
        'total_collection': total_collection,
        'total_loans': total_loans,
        'pending_dues': pending_dues,
        'monthly_summary_json': json.dumps(monthly_summary),
        'recent_payments': recent_payments,
        'current_year': current_year,
    }
    return render(request, 'admin_dashboard.html', context)

@login_required
@member_required
def member_dashboard(request):
    member = request.user.member_profile
    total_paid = member.total_paid
    loans = member.loans.filter(status='Active')
    total_loan_due = loans.aggregate(total=Sum('remaining_due'))['total'] or 0
    payments = member.payments.filter(status='Paid').order_by('-payment_date')[:10]

    context = {
        'member': member,
        'total_paid': total_paid,
        'total_loan_due': total_loan_due,
        'active_loans': loans,
        'recent_payments': payments,
    }
    return render(request, 'member_dashboard.html', context)

# ---------------------- Member Management (Admin) ----------------------
@login_required
@admin_required
def member_list(request):
    members = Member.objects.filter(role='member')
    search = request.GET.get('search')
    if search:
        members = members.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(phone__icontains=search)
        )
    return render(request, 'member_list.html', {'members': members})

@login_required
@admin_required
def member_create(request):
    if request.method == 'POST':
        form = MemberForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Member {user.username} created successfully')
            return redirect('member_list')
    else:
        form = MemberForm()
    return render(request, 'member_form.html', {'form': form, 'title': 'Add Member'})

@login_required
@admin_required
def member_edit(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        form = MemberForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, 'Member updated successfully')
            return redirect('member_list')
    else:
        form = MemberForm(instance=member)
    return render(request, 'member_form.html', {'form': form, 'title': 'Edit Member'})

@login_required
@admin_required
def member_delete(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        member.user.delete()
        messages.success(request, 'Member deleted successfully')
        return redirect('member_list')
    return render(request, 'confirm_delete.html', {'object': member})

# ---------------------- Pedi Management ----------------------
@login_required
@admin_required
def pedi_list(request):
    pedis = Pedi.objects.all()
    return render(request, 'pedi_list.html', {'pedis': pedis})

@login_required
@admin_required
def pedi_create(request):
    if request.method == 'POST':
        form = PediForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pedi created successfully')
            return redirect('pedi_list')
    else:
        form = PediForm()
    return render(request, 'pedi_form.html', {'form': form, 'title': 'Create Pedi'})

@login_required
@admin_required
def pedi_edit(request, pk):
    pedi = get_object_or_404(Pedi, pk=pk)
    if request.method == 'POST':
        form = PediForm(request.POST, instance=pedi)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pedi updated successfully')
            return redirect('pedi_list')
    else:
        form = PediForm(instance=pedi)
    return render(request, 'pedi_form.html', {'form': form, 'title': 'Edit Pedi'})

@login_required
@admin_required
def assign_members(request, pedi_id):
    pedi = get_object_or_404(Pedi, pk=pedi_id)
    members = Member.objects.filter(role='member', is_active=True)
    assigned = MemberPedi.objects.filter(pedi=pedi).values_list('member_id', flat=True)

    if request.method == 'POST':
        selected_members = request.POST.getlist('members')
        MemberPedi.objects.filter(pedi=pedi).delete()
        for member_id in selected_members:
            member = Member.objects.get(pk=member_id)
            MemberPedi.objects.create(member=member, pedi=pedi)
            for month in range(1, pedi.duration_months + 1):
                Payment.objects.get_or_create(
                    member=member,
                    pedi=pedi,
                    month=month,
                    year=pedi.start_date.year,
                    defaults={'amount': pedi.monthly_amount, 'status': 'Pending'}
                )
        messages.success(request, 'Members assigned successfully')
        return redirect('pedi_list')

    context = {'pedi': pedi, 'members': members, 'assigned': assigned}
    return render(request, 'assign_members.html', context)

# ---------------------- Monthly Payments ----------------------
@login_required
@admin_required
def monthly_payments(request, pedi_id=None):
    pedis = Pedi.objects.filter(is_active=True)
    selected_pedi = None
    payments = []

    if pedi_id:
        selected_pedi = get_object_or_404(Pedi, pk=pedi_id)
        members_in_pedi = MemberPedi.objects.filter(pedi=selected_pedi, status='Active').select_related('member')

        current_month = int(request.GET.get('month', timezone.now().month))
        current_year = int(request.GET.get('year', timezone.now().year))

        for mp in members_in_pedi:
            payment, created = Payment.objects.get_or_create(
                member=mp.member,
                pedi=selected_pedi,
                month=current_month,
                year=current_year,
                defaults={'amount': selected_pedi.monthly_amount, 'status': 'Pending'}
            )
            payments.append({
                'member': mp.member,
                'payment': payment,
                'amount': payment.amount,
                'status': payment.status,
            })

        if request.method == 'POST':
            for item in payments:
                if request.POST.get(f'payment_{item["payment"].id}'):
                    payment = item['payment']
                    payment.status = 'Paid'
                    payment.payment_date = timezone.now()
                    payment.payment_method = 'Cash'
                    payment.save()
            messages.success(request, 'Payments updated successfully')
            return redirect('monthly_payments', pedi_id=selected_pedi.id)

    context = {
        'pedis': pedis,
        'selected_pedi': selected_pedi,
        'payments': payments,
        'current_month': timezone.now().month,
        'current_year': timezone.now().year,
        'months': range(1, 13),
    }
    return render(request, 'monthly_payments.html', context)

# ---------------------- Loan Management ----------------------
@login_required
@admin_required
def loan_list(request):
    loans = Loan.objects.all().select_related('member')
    status_filter = request.GET.get('status')
    if status_filter:
        loans = loans.filter(status=status_filter)
    return render(request, 'loan_list.html', {'loans': loans})

@login_required
@admin_required
def loan_create(request):
    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Loan issued successfully')
            return redirect('loan_list')
    else:
        form = LoanForm()
    return render(request, 'loan_form.html', {'form': form, 'title': 'Issue Loan'})

@login_required
@admin_required
def loan_edit(request, pk):
    loan = get_object_or_404(Loan, pk=pk)
    if request.method == 'POST':
        form = LoanForm(request.POST, instance=loan)
        if form.is_valid():
            form.save()
            messages.success(request, 'Loan updated successfully')
            return redirect('loan_list')
    else:
        form = LoanForm(instance=loan)
    return render(request, 'loan_form.html', {'form': form, 'title': 'Edit Loan'})

# ---------------------- Member Views for Loans & Payments ----------------------
@login_required
@member_required
def member_loans(request):
    member = request.user.member_profile
    loans = member.loans.all()
    return render(request, 'member_loans.html', {'loans': loans})

@login_required
@member_required
def member_payments(request):
    member = request.user.member_profile
    payments = member.payments.all().order_by('-year', '-month')
    return render(request, 'member_payments.html', {'payments': payments})

@login_required
@member_required
def payment_history(request):
    member = request.user.member_profile
    payments = member.payments.filter(status='Paid').order_by('-payment_date')
    return render(request, 'payment_history.html', {'payments': payments})

# ---------------------- loan Payment ----------------------
@login_required
@member_required
def loan_pay_online(request, loan_id):
    loan = get_object_or_404(Loan, pk=loan_id, member=request.user.member_profile)
    if loan.status == 'Closed':
        messages.warning(request, 'This loan is already closed.')
        return redirect('member_loans')

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', 0))
        if amount <= 0:
            messages.error(request, 'Please enter a valid amount.')
            return redirect('loan_pay_online', loan_id=loan.id)
        if amount > loan.remaining_due:
            messages.error(request, f'Amount cannot exceed remaining due ({loan.remaining_due}).')
            return redirect('loan_pay_online', loan_id=loan.id)

        # Create Razorpay order
        order_amount = int(amount * 100)
        try:
            razorpay_order = client.order.create({
                'amount': order_amount,
                'currency': 'INR',
                'payment_capture': '1',
                'notes': {
                    'loan_id': loan.id,
                    'member_id': loan.member.id,
                    'amount': str(amount)
                }
            })
        except Exception as e:
            messages.error(request, f'Payment gateway error: {str(e)}')
            return redirect('member_loans')

        # Save transaction
        LoanTransaction.objects.create(
            loan=loan,
            amount=amount,
            razorpay_order_id=razorpay_order['id'],
            status='Created'
        )

        # Build JSON data for template
        payment_data = {
            'razorpay_key_id': RAZORPAY_KEY_ID,
            'amount_paise': order_amount,
            'razorpay_order_id': razorpay_order['id'],
            'csrf_token': get_token(request),
            'success_url': request.build_absolute_uri(reverse('loan_payment_online_success')),
        }
        context = {
            'loan': loan,
            'loan_payment_data': payment_data,
        }
        return render(request, 'loan_payment_gateway.html', context)

    # GET request – show payment form
    return render(request, 'loan_pay_online.html', {'loan': loan})

@login_required
def loan_payment_online_success(request):
    if request.method == 'POST':
        razorpay_payment_id = request.POST.get('razorpay_payment_id')
        razorpay_order_id = request.POST.get('razorpay_order_id')
        razorpay_signature = request.POST.get('razorpay_signature')

        params_dict = {
            'razorpay_order_id': razorpay_order_id,
            'razorpay_payment_id': razorpay_payment_id,
            'razorpay_signature': razorpay_signature
        }

        try:
            client.utility.verify_payment_signature(params_dict)
            loan_trans = LoanTransaction.objects.get(razorpay_order_id=razorpay_order_id)
            loan_trans.razorpay_payment_id = razorpay_payment_id
            loan_trans.razorpay_signature = razorpay_signature
            loan_trans.status = 'Success'
            loan_trans.save()

            # Create LoanPayment record
            LoanPayment.objects.create(
                loan=loan_trans.loan,
                amount=loan_trans.amount,
                transaction_id=razorpay_payment_id,
                payment_method='Online'
            )
            messages.success(request, f'Loan payment of ₹{loan_trans.amount} successful!')
            return redirect('member_loans')
        except Exception as e:
            messages.error(request, f'Payment verification failed: {str(e)}')
            return redirect('member_loans')
    return redirect('member_loans')

@login_required
@member_required
def loan_payment_history(request):
    member = request.user.member_profile
    payments = LoanPayment.objects.filter(loan__member=member).order_by('-payment_date')
    return render(request, 'loan_payment_history.html', {'payments': payments})

@login_required
@admin_required
def admin_loan_payments(request):
    payments = LoanPayment.objects.select_related('loan__member__user').order_by('-payment_date')
    return render(request, 'admin_loan_payments.html', {'payments': payments})
# ---------------------- Online Payment ----------------------

@login_required
@member_required
def make_payment(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id, member=request.user.member_profile)
    if payment.status == 'Paid':
        messages.warning(request, 'This payment is already completed')
        return redirect('member_payments')

    if request.method == 'POST':
        # Create Razorpay Order (amount in paise)
        order_amount = int(payment.amount * 100)  # convert to paise (integer)
        order_currency = 'INR'
        razorpay_order = client.order.create({
            'amount': order_amount,
            'currency': order_currency,
            'payment_capture': '1',
            'notes': {
                'payment_id': payment.id,
                'member_id': payment.member.id
            }
        })

        # Save transaction
        Transaction.objects.create(
            member=payment.member,
            payment=payment,
            razorpay_order_id=razorpay_order['id'],
            amount=payment.amount,
            status='Created'
        )

        payment_data = {
            'razorpay_key_id': RAZORPAY_KEY_ID,
            'amount_paise': order_amount,
            'razorpay_order_id': razorpay_order['id'],
            'csrf_token': get_token(request),
            'success_url': request.build_absolute_uri(reverse('payment_success')),
            'pedi_name': payment.pedi.name if payment.pedi else '',
            'month': payment.month,
            'year': payment.year,
        }
        context = {
            'payment': payment,
            'payment_data_json': payment_data,
            'razorpay_order_id': razorpay_order['id'],
            'razorpay_key_id': RAZORPAY_KEY_ID,
            'amount': payment.amount,           # for display
            'amount_paise': order_amount,       # for Razorpay (integer paise)
        }
        return render(request, 'payment_gateway.html', context)

    return render(request, 'make_payment.html', {'payment': payment})

@login_required
def payment_success(request):
    if request.method == 'POST':
        razorpay_payment_id = request.POST.get('razorpay_payment_id')
        razorpay_order_id = request.POST.get('razorpay_order_id')
        razorpay_signature = request.POST.get('razorpay_signature')

        params_dict = {
            'razorpay_order_id': razorpay_order_id,
            'razorpay_payment_id': razorpay_payment_id,
            'razorpay_signature': razorpay_signature
        }

        try:
            client.utility.verify_payment_signature(params_dict)
            transaction = Transaction.objects.get(razorpay_order_id=razorpay_order_id)
            transaction.razorpay_payment_id = razorpay_payment_id
            transaction.razorpay_signature = razorpay_signature
            transaction.status = 'Success'
            transaction.save()

            payment = transaction.payment
            payment.status = 'Paid'
            payment.payment_date = timezone.now()
            payment.payment_method = 'Online'
            payment.transaction_id = razorpay_payment_id
            payment.razorpay_order_id = razorpay_order_id
            payment.razorpay_payment_id = razorpay_payment_id
            payment.save()

            messages.success(request, 'Payment successful!')
            return redirect('member_payments')
        except Exception:
            messages.error(request, 'Payment verification failed')
            return redirect('member_payments')

    return redirect('member_dashboard')

# ---------------------- Reports & Export ----------------------
@login_required
@admin_required
def reports(request):
    return render(request, 'reports.html')

@login_required
@admin_required
def export_members_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    headers = ['Name', 'Username', 'Phone', 'Address', 'Joined Date', 'Total Paid']
    ws.append(headers)
    members = Member.objects.filter(role='member')
    for member in members:
        ws.append([
            member.user.get_full_name(),
            member.user.username,
            member.phone,
            member.address,
            member.joined_date.strftime('%Y-%m-%d'),
            float(member.total_paid)
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    wb.save(response)
    return response

@login_required
@admin_required
def export_payments_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ['Member', 'Pedi', 'Month', 'Year', 'Amount', 'Status', 'Payment Date']
    ws.append(headers)
    payments = Payment.objects.select_related('member', 'pedi').all()
    for payment in payments:
        ws.append([
            payment.member.user.get_full_name(),
            payment.pedi.name,
            payment.month,
            payment.year,
            float(payment.amount),
            payment.status,
            payment.payment_date.strftime('%Y-%m-%d %H:%M') if payment.payment_date else ''
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
    wb.save(response)
    return response

@login_required
@admin_required
def export_loans_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Loans"
    headers = ['Member', 'Amount', 'Interest Rate', 'Total Payable', 'Paid Amount', 'Remaining Due', 'Status']
    ws.append(headers)
    loans = Loan.objects.select_related('member').all()
    for loan in loans:
        ws.append([
            loan.member.user.get_full_name(),
            float(loan.amount),
            float(loan.interest_rate),
            float(loan.total_payable),
            float(loan.paid_amount),
            float(loan.remaining_due),
            loan.status
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=loans.xlsx'
    wb.save(response)
    return response


@login_required
@member_required
def apply_loan(request):
    settings = LoanApplicationSettings.objects.first()
    if not settings:
        messages.error(request, 'Loan application period not configured.')
        return redirect('member_dashboard')

    today = timezone.now().date()
    if today < settings.start_date or today > settings.end_date:
        messages.warning(request, 'Loan applications are currently closed.')
        return redirect('member_dashboard')

    # --- NEW: Check for existing active loan ---
    member = request.user.member_profile
    if Loan.objects.filter(member=member, status='Active').exists():
        messages.warning(request, 'You already have an active loan. You cannot apply for a new loan until the current loan is fully paid.')
        return redirect('member_dashboard')
    # -------------------------------------------

    # Check if member already has a pending application
    if LoanApplication.objects.filter(member=member, status='Pending').exists():
        messages.error(request, 'You already have a pending loan application.')
        return redirect('member_dashboard')

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', 0))
        purpose = request.POST.get('purpose', '')
        if amount <= 0:
            messages.error(request, 'Please enter a valid amount.')
            return render(request, 'apply_loan.html', {'settings': settings})

        LoanApplication.objects.create(
            member=member,
            requested_amount=amount,
            purpose=purpose,
            status='Pending'
        )
        messages.success(request, 'Your loan application has been submitted.')
        return redirect('member_dashboard')

    return render(request, 'apply_loan.html', {'settings': settings})

@login_required
@admin_required
def admin_loan_applications(request):
    applications = LoanApplication.objects.select_related('member__user').order_by('-applied_date')
    status_filter = request.GET.get('status')
    if status_filter:
        applications = applications.filter(status=status_filter)
    return render(request, 'admin_loan_applications.html', {'applications': applications})

@login_required
@admin_required
def approve_loan_application(request, pk):
    application = get_object_or_404(LoanApplication, pk=pk)
    if application.status != 'Pending':
        messages.warning(request, 'This application is no longer pending.')
        return redirect('admin_loan_applications')

    settings = LoanApplicationSettings.objects.first()
    if request.method == 'POST':
        interest_rate = Decimal(request.POST.get('interest_rate', settings.default_interest_rate))
        due_date = request.POST.get('due_date')
        if not due_date:
            due_date = (timezone.now().date() + relativedelta(months=settings.default_loan_duration_months))
        else:
            from datetime import datetime
            due_date = datetime.strptime(due_date, '%Y-%m-%d').date()

        # Create actual Loan
        loan = Loan.objects.create(
            member=application.member,
            amount=application.requested_amount,
            interest_rate=interest_rate,
            due_date=due_date,
            status='Active'
        )
        # Update application
        application.status = 'Approved'
        application.approved_date = timezone.now()
        application.approved_interest_rate = interest_rate
        application.approved_due_date = due_date
        application.admin_remarks = request.POST.get('remarks', '')
        application.save()

        messages.success(request, f'Loan application approved. Loan #{loan.id} created.')
        return redirect('admin_loan_applications')

    # Pre-fill form with defaults
    default_due = timezone.now().date() + relativedelta(months=settings.default_loan_duration_months)
    context = {
        'application': application,
        'default_interest': settings.default_interest_rate,
        'default_due': default_due,
    }
    return render(request, 'approve_loan_application.html', context)

@login_required
@admin_required
def reject_loan_application(request, pk):
    application = get_object_or_404(LoanApplication, pk=pk)
    if application.status != 'Pending':
        messages.warning(request, 'This application is no longer pending.')
        return redirect('admin_loan_applications')
    application.status = 'Rejected'
    application.save()
    messages.success(request, 'Loan application rejected.')
    return redirect('admin_loan_applications')

# admin loan setting view
@login_required
@admin_required
def admin_loan_settings(request):
    settings = LoanApplicationSettings.objects.first()
    if not settings:
        settings = LoanApplicationSettings.objects.create(
            start_date=timezone.now().date(),
            end_date=timezone.now().date() + timedelta(days=30),
            default_interest_rate=10.0,
            default_loan_duration_months=12
        )

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        interest_rate = request.POST.get('default_interest_rate')
        duration = request.POST.get('default_loan_duration_months')
        if start_date and end_date and interest_rate and duration:
            from datetime import datetime
            settings.start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            settings.end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            settings.default_interest_rate = Decimal(interest_rate)
            settings.default_loan_duration_months = int(duration)
            settings.save()
            messages.success(request, 'Loan application settings updated.')
        else:
            messages.error(request, 'Please fill all fields.')
        return redirect('admin_loan_settings')

    context = {'settings': settings}
    return render(request, 'admin_loan_settings.html', context)